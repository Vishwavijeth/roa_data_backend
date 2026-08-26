from datetime import date
import io
import re
from typing import Dict, List, Optional

import pandas as pd
from fastapi import APIRouter, Depends, Query, Response
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, case, distinct, func, or_, select
from sqlalchemy.orm import Session

from db import get_db
from models.skyslope.meta import Office, Stage
from models.skyslope.property import SaleProperty
from models.skyslope.sale import Sale
from models.skyslope.users import SkyslopeUser
from services.states import STATE_FULL_NAME_MAP

router = APIRouter()


STATE_CODE_BY_FULL_NAME = {
    full_name.upper(): state_code
    for state_code, full_name in STATE_FULL_NAME_MAP.items()
}


# ============================================================
# COMMON ORM EXPRESSIONS / HELPERS
# ============================================================

def reviewer_name_expr():
    full_name = func.nullif(
        func.trim(
            func.concat_ws(
                " ",
                SkyslopeUser.firstname,
                SkyslopeUser.lastname,
            )
        ),
        "",
    )

    return case(
        # No reviewer GUID on the sale itself.
        (Sale.reviewerguid.is_(None), "Unassigned"),

        # Reviewer GUID exists on the sale, but there is no matching users row.
        (SkyslopeUser.userguid.is_(None), "No User Record"),

        # Matching users row exists, but firstname + lastname is empty/null.
        (full_name.is_(None), "No User Record"),

        else_=full_name,
    )


def normalized_status_expr():
    return func.lower(func.trim(func.coalesce(Sale.status, "")))


def normalized_state_expr():
    return func.coalesce(
        func.nullif(func.upper(func.trim(SaleProperty.state)), ""),
        "UNKNOWN",
    )


def parse_date(value: Optional[str]) -> Optional[date]:
    if not value:
        return None
    return date.fromisoformat(value)


def status_field_name(status: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "_", status.strip().lower()).strip("_")
    return f"transactions_{normalized}"


def get_distinct_statuses(db: Session) -> List[str]:
    status_value = func.trim(Sale.status)

    stmt = (
        select(status_value)
        .where(
            Sale.status.is_not(None),
            func.trim(Sale.status) != "",
        )
        .distinct()
    )

    statuses = list(db.scalars(stmt).all())
    return sorted(statuses, key=str.casefold)


def build_status_field_map(statuses: List[str]) -> Dict[str, str]:
    field_map: Dict[str, str] = {}
    used_fields: Dict[str, str] = {}

    for status in statuses:
        base_field = status_field_name(status)
        field_name = base_field
        counter = 2

        while field_name in used_fields and used_fields[field_name].casefold() != status.casefold():
            field_name = f"{base_field}_{counter}"
            counter += 1

        field_map[status] = field_name
        used_fields[field_name] = status

    return field_map


def apply_common_filters_orm(
    stmt,
    *,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    state: Optional[List[str]] = None,
    stage_name: Optional[List[str]] = None,
    status: Optional[List[str]] = None,
    reviewer: Optional[List[str]] = None,
    type_of_sale: Optional[List[str]] = None,
):
    conditions = []

    parsed_from_date = parse_date(from_date)
    parsed_to_date = parse_date(to_date)

    if parsed_from_date:
        conditions.append(func.date(Sale.escrowclosingdate) >= parsed_from_date)

    if parsed_to_date:
        conditions.append(func.date(Sale.escrowclosingdate) <= parsed_to_date)

    # State is selected from sale_property.state, but filtering is done
    # against the transaction's office name.
    # Example: TX -> office name starts with "TX" or "Texas".
    if state:
        office_state_conditions = []

        for value in state:
            if not value or not value.strip():
                continue

            selected_state = value.strip().upper()
            state_code = STATE_CODE_BY_FULL_NAME.get(selected_state, selected_state)
            full_state_name = STATE_FULL_NAME_MAP.get(state_code)

            if not full_state_name:
                continue

            office_state_conditions.append(
                or_(
                    Office.officename.ilike(f"{state_code}%"),
                    Office.officename.ilike(f"{full_state_name}%"),
                )
            )

        if office_state_conditions:
            conditions.append(or_(*office_state_conditions))

    if stage_name:
        normalized_stages = [
            value.strip().lower()
            for value in stage_name
            if value and value.strip()
        ]

        if normalized_stages:
            conditions.append(
                func.lower(func.trim(Stage.name)).in_(normalized_stages)
            )

    if status:
        normalized_statuses = [
            value.strip().lower()
            for value in status
            if value and value.strip()
        ]

        if normalized_statuses:
            conditions.append(
                normalized_status_expr().in_(normalized_statuses)
            )

    if reviewer:
        reviewer_values = [
            value.strip()
            for value in reviewer
            if value and value.strip()
        ]

        if reviewer_values:
            conditions.append(
                reviewer_name_expr().in_(reviewer_values)
            )

    if type_of_sale:
        normalized_sale_types = [
            value.strip().lower()
            for value in type_of_sale
            if value and value.strip()
        ]

        if normalized_sale_types:
            conditions.append(
                func.lower(func.trim(Sale.dealtype)).in_(normalized_sale_types)
            )

    if conditions:
        stmt = stmt.where(and_(*conditions))

    return stmt


def status_count_columns(
    statuses: List[str],
    status_field_map: Dict[str, str],
):
    status_expr = normalized_status_expr()
    columns = []

    for status in statuses:
        columns.append(
            func.count(distinct(Sale.saleguid))
            .filter(status_expr == status.strip().lower())
            .label(status_field_map[status])
        )

    columns.append(
        func.count(distinct(Sale.saleguid)).label("total_transactions")
    )

    return columns


def build_reviewer_dashboard_stmt(
    *,
    statuses: List[str],
    status_field_map: Dict[str, str],
    from_date: Optional[str],
    to_date: Optional[str],
    state: Optional[List[str]],
    stage_name: Optional[List[str]],
    status: Optional[List[str]],
    reviewer: Optional[List[str]],
    type_of_sale: Optional[List[str]],
    include_reviewer_guid: bool,
):
    reviewer_full_name = reviewer_name_expr().label("reviewer_full_name")

    columns = []

    if include_reviewer_guid:
        columns.append(Sale.reviewerguid.label("reviewerguid"))

    columns.extend(
        [
            reviewer_full_name,
            *status_count_columns(statuses, status_field_map),
        ]
    )

    stmt = (
        select(*columns)
        .select_from(Sale)
        .outerjoin(
            SkyslopeUser,
            Sale.reviewerguid == SkyslopeUser.userguid,
        )
        .outerjoin(
            Stage,
            Sale.stageid == Stage.stageid,
        )
        .outerjoin(
            Office,
            Sale.officeguid == Office.officeguid,
        )
    )

    stmt = apply_common_filters_orm(
        stmt,
        from_date=from_date,
        to_date=to_date,
        state=state,
        stage_name=stage_name,
        status=status,
        reviewer=reviewer,
        type_of_sale=type_of_sale,
    )

    if include_reviewer_guid:
        stmt = stmt.group_by(
            Sale.reviewerguid,
            reviewer_name_expr(),
        )
    else:
        stmt = stmt.group_by(reviewer_name_expr())

    return stmt.order_by(reviewer_name_expr())


def get_status_count(row: dict, status: str, status_field_map: Dict[str, str]) -> int:
    for existing_status, field_name in status_field_map.items():
        if existing_status.casefold() == status.casefold():
            return row.get(field_name, 0) or 0
    return 0


def create_export_rows(
    rows: List[dict],
    statuses: List[str],
    status_field_map: Dict[str, str],
    first_column_name: str,
    first_column_key: str,
) -> List[dict]:
    export_rows = []

    for row in rows:
        export_row = {
            first_column_name: row.get(first_column_key) or "",
        }

        for status in statuses:
            export_row[status] = row.get(status_field_map[status], 0) or 0

        export_row["Total Transactions"] = row.get("total_transactions", 0) or 0
        export_rows.append(export_row)

    return export_rows


def style_standard_dashboard_sheet(worksheet, df: pd.DataFrame):
    worksheet.freeze_panes = "A2"

    font_header = Font(name="Segoe UI", size=11, bold=True)
    align_header = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    font_body = Font(name="Segoe UI", size=10)

    worksheet.row_dimensions[1].height = 28

    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = font_header
        cell.alignment = align_header

    for row_num in range(2, len(df) + 2):
        worksheet.row_dimensions[row_num].height = 20

        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = font_body

            if col_num == 1:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                )
            else:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

    for col in worksheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        worksheet.column_dimensions[col_letter].width = max(max_len + 3, 14)


# ============================================================
# REVIEWER DASHBOARD
# ============================================================

@router.get("/reviewer-dashboard")
def reviewer_dashboard(
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    state: Optional[List[str]] = Query(None),
    stage_name: Optional[List[str]] = Query(None),
    status: Optional[List[str]] = Query(None),
    reviewer: Optional[List[str]] = Query(None),
    type_of_sale: Optional[List[str]] = Query(None),
    db: Session = Depends(get_db),
):
    statuses = get_distinct_statuses(db)
    status_field_map = build_status_field_map(statuses)

    stmt = build_reviewer_dashboard_stmt(
        statuses=statuses,
        status_field_map=status_field_map,
        from_date=from_date,
        to_date=to_date,
        state=state,
        stage_name=stage_name,
        status=status,
        reviewer=reviewer,
        type_of_sale=type_of_sale,
        include_reviewer_guid=True,
    )

    rows = [
        dict(row)
        for row in db.execute(stmt).mappings().all()
    ]

    summary = {
        "count": len(rows),
        "outstanding_transactions": sum(
            get_status_count(row, "pending", status_field_map)
            + get_status_count(row, "expired", status_field_map)
            for row in rows
        ),
        "closed_transactions": sum(
            get_status_count(row, "closed", status_field_map)
            + get_status_count(row, "archived", status_field_map)
            for row in rows
        ),
    }

    return {
        "summary": summary,
        "status_fields": {
            status: status_field_map[status]
            for status in statuses
        },
        "data": rows,
    }


# ============================================================
# REVIEWER DASHBOARD DOWNLOAD
# ============================================================

@router.get("/reviewer-dashboard/download")
def download_reviewer_dashboard(
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    state: Optional[List[str]] = Query(None),
    stage_name: Optional[List[str]] = Query(None),
    status: Optional[List[str]] = Query(None),
    reviewer: Optional[List[str]] = Query(None),
    type_of_sale: Optional[List[str]] = Query(None),
    db: Session = Depends(get_db),
):
    statuses = get_distinct_statuses(db)
    status_field_map = build_status_field_map(statuses)

    stmt = build_reviewer_dashboard_stmt(
        statuses=statuses,
        status_field_map=status_field_map,
        from_date=from_date,
        to_date=to_date,
        state=state,
        stage_name=stage_name,
        status=status,
        reviewer=reviewer,
        type_of_sale=type_of_sale,
        include_reviewer_guid=False,
    )

    rows = [
        dict(row)
        for row in db.execute(stmt).mappings().all()
    ]

    rows_to_export = create_export_rows(
        rows=rows,
        statuses=statuses,
        status_field_map=status_field_map,
        first_column_name="Reviewer Name",
        first_column_key="reviewer_full_name",
    )

    export_columns = [
        "Reviewer Name",
        *statuses,
        "Total Transactions",
    ]

    df = pd.DataFrame(rows_to_export, columns=export_columns)

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(
            writer,
            sheet_name="Reviewer Dashboard",
            index=False,
        )

        worksheet = writer.sheets["Reviewer Dashboard"]
        style_standard_dashboard_sheet(worksheet, df)

    output.seek(0)

    filename = "reviewer_dashboard_report.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


# ============================================================
# REVIEWER FILTERS
# ============================================================

@router.get("/reviewers/filters")
def reviewer_dashboard_filters(
    db: Session = Depends(get_db),
):
    stage_stmt = (
        select(Stage.name)
        .where(
            Stage.name.is_not(None),
            func.trim(Stage.name) != "",
        )
        .distinct()
        .order_by(Stage.name)
    )

    state_value = func.upper(func.trim(SaleProperty.state))

    state_stmt = (
        select(state_value)
        .where(
            SaleProperty.state.is_not(None),
            func.trim(SaleProperty.state) != "",
        )
        .distinct()
        .order_by(state_value)
    )

    reviewer_value = reviewer_name_expr().label("reviewer_name")

    reviewer_stmt = (
        select(reviewer_value)
        .select_from(Sale)
        .outerjoin(
            SkyslopeUser,
            Sale.reviewerguid == SkyslopeUser.userguid,
        )
        .distinct()
        .order_by(reviewer_value)
    )

    type_of_sale_stmt = (
        select(Sale.dealtype)
        .where(
            Sale.dealtype.is_not(None),
            func.trim(Sale.dealtype) != "",
        )
        .distinct()
        .order_by(Sale.dealtype)
    )

    stage_list = list(db.scalars(stage_stmt).all())
    state_list = list(db.scalars(state_stmt).all())
    status_list = get_distinct_statuses(db)
    reviewer_list = list(db.scalars(reviewer_stmt).all())
    type_of_sale_list = list(db.scalars(type_of_sale_stmt).all())

    return {
        "stage_list": stage_list,
        "state_list": state_list,
        "status_list": status_list,
        "reviewer_list": reviewer_list,
        "type_of_sale_list": type_of_sale_list,
    }


# ============================================================
# UNASSIGNED REVIEWER REPORT
# ============================================================

@router.get("/reviewer-dashboard/unassigned/download")
def download_unassigned_reviewer_state_report(
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    state: Optional[List[str]] = Query(None),
    stage_name: Optional[List[str]] = Query(None),
    status: Optional[List[str]] = Query(None),
    reviewer: Optional[List[str]] = Query(None),
    type_of_sale: Optional[List[str]] = Query(None),
    db: Session = Depends(get_db),
):
    statuses = get_distinct_statuses(db)
    status_field_map = build_status_field_map(statuses)
    state_value = normalized_state_expr().label("state")

    stmt = (
        select(
            state_value,
            *status_count_columns(statuses, status_field_map),
        )
        .select_from(Sale)
        .outerjoin(
            SkyslopeUser,
            Sale.reviewerguid == SkyslopeUser.userguid,
        )
        .outerjoin(
            SaleProperty,
            Sale.saleguid == SaleProperty.saleguid,
        )
        .outerjoin(
            Stage,
            Sale.stageid == Stage.stageid,
        )
        .outerjoin(
            Office,
            Sale.officeguid == Office.officeguid,
        )
        .where(Sale.reviewerguid.is_(None))
    )

    stmt = apply_common_filters_orm(
        stmt,
        from_date=from_date,
        to_date=to_date,
        state=state,
        stage_name=stage_name,
        status=status,
        reviewer=None,
        type_of_sale=type_of_sale,
    )

    stmt = (
        stmt
        .group_by(normalized_state_expr())
        .order_by(normalized_state_expr())
    )

    rows = [
        dict(row)
        for row in db.execute(stmt).mappings().all()
    ]

    applied_filters = []

    if from_date:
        applied_filters.append(f"From Date: {from_date}")

    if to_date:
        applied_filters.append(f"To Date: {to_date}")

    if state:
        applied_filters.append(f"State: {', '.join(state)}")

    if stage_name:
        applied_filters.append(f"Stage: {', '.join(stage_name)}")

    if status:
        applied_filters.append(f"Status: {', '.join(status)}")

    if type_of_sale:
        applied_filters.append(f"Type of Sale: {', '.join(type_of_sale)}")

    applied_filters.append("Reviewer: Unassigned")

    rows_to_export = create_export_rows(
        rows=rows,
        statuses=statuses,
        status_field_map=status_field_map,
        first_column_name="State",
        first_column_key="state",
    )

    export_columns = [
        "State",
        *statuses,
        "Total Transactions",
    ]

    df = pd.DataFrame(rows_to_export, columns=export_columns)

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_name = "Unassigned by State"

        df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            startrow=0,
            startcol=1,
        )

        worksheet = writer.sheets[sheet_name]

        font_header = Font(name="Segoe UI", size=11, bold=True)
        font_body = Font(name="Segoe UI", size=10)

        worksheet["A1"] = "Filters"
        worksheet["A1"].font = font_header
        worksheet["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        for idx, filter_text in enumerate(applied_filters, start=2):
            cell = worksheet.cell(row=idx, column=1)
            cell.value = filter_text
            cell.font = font_body
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )

        for col_num in range(2, len(df.columns) + 2):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = font_header
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for row_num in range(2, len(df) + 2):
            for col_num in range(2, len(df.columns) + 2):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = font_body

                if col_num == 2:
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                    )
                else:
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                    )

        worksheet.row_dimensions[1].height = 28
        max_rows = max(len(applied_filters) + 1, len(df) + 1)

        for row_num in range(2, max_rows + 1):
            worksheet.row_dimensions[row_num].height = 20

        worksheet.freeze_panes = "B2"

        last_col_letter = get_column_letter(len(df.columns) + 1)
        last_row = len(df) + 1 if len(df) > 0 else 1
        worksheet.auto_filter.ref = f"B1:{last_col_letter}{last_row}"

        worksheet.column_dimensions["A"].width = 30

        for col_num in range(2, len(df.columns) + 2):
            col_letter = get_column_letter(col_num)
            max_len = 0

            for row_num in range(1, len(df) + 2):
                value = worksheet.cell(row=row_num, column=col_num).value

                if value is not None:
                    max_len = max(max_len, len(str(value)))

            if worksheet.cell(row=1, column=col_num).value == "State":
                worksheet.column_dimensions[col_letter].width = 14
            else:
                worksheet.column_dimensions[col_letter].width = max(
                    max_len + 3,
                    14,
                )

    output.seek(0)

    filename = "unassigned_reviewer_state_report.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )