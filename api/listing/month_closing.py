from fastapi import APIRouter, Query, Response, Depends
from typing import List, Optional
from sqlalchemy import text
from sqlalchemy.orm import Session
from db import get_db
import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal
import datetime


router = APIRouter()


def fetch_month_closing_data(
    status: str = "all",
    skyslope: bool = False,
    state: Optional[List[str]] = None,
    from_close_date: str = None,
    to_close_date: str = None,
    transaction_specialist: Optional[List[str]] = None,
    search: str = None,
    mismatch: bool = False,
    pending_subfilter: Optional[List[str]] = None,
    page: int = None,
    page_size: int = None,
    db: Session = None
):
    try:
        state = state or []
        transaction_specialist = transaction_specialist or []
        pending_subfilter = pending_subfilter or []

        state_list = [v.strip() for s in state for v in s.split(",") if v.strip()]
        ts_list = [v.strip() for s in transaction_specialist for v in s.split(",") if v.strip()]
        ps_list = [v.strip() for s in pending_subfilter for v in s.split(",") if v.strip()]

        if skyslope:
            params = {}
            where_clause = """
                WHERE 1=1
                  AND NOT EXISTS (
                      SELECT 1
                      FROM brokerage_engine be2
                      WHERE be2.skyslopefileid = s.saleguid
                  )
                  AND NOT EXISTS (
                      SELECT 1
                      FROM otherincome_transactions oit2
                      WHERE oit2.skyslopefileid = s.saleguid
                  )
            """

            if status != "all":
                cleaned_status = status.strip().lower()
                if cleaned_status == "pending":
                    where_clause += """
                        AND (
                            LOWER(TRIM(COALESCE(s.status, ''))) = 'pending'
                            OR LOWER(TRIM(COALESCE(s.status, ''))) = 'active'
                            OR LOWER(TRIM(COALESCE(s.status, ''))) = 'in_progress'
                        )
                    """
                elif cleaned_status == "closed":
                    where_clause += """
                        AND LOWER(TRIM(COALESCE(s.status, ''))) = 'closed'
                    """
                elif cleaned_status == "cancelled":
                    where_clause += """
                        AND LOWER(TRIM(COALESCE(s.status, ''))) IN ('cancelled', 'canceled', 'canceled/app', 'canceled/pend')
                    """
                else:
                    where_clause += """
                        AND LOWER(TRIM(COALESCE(s.status, ''))) NOT IN (
                            'pending', 'active', 'in_progress', 'closed',
                            'cancelled', 'canceled', 'canceled/app', 'canceled/pend'
                        )
                    """

            if state_list:
                placeholders = ", ".join(f":state_{i}" for i in range(len(state_list)))
                where_clause += f" AND LOWER(sp.state) IN ({placeholders})"
                for i, v in enumerate(state_list):
                    params[f"state_{i}"] = v.lower()

            if from_close_date:
                where_clause += " AND s.escrowclosingdate >= :from_close_date"
                params["from_close_date"] = from_close_date

            if to_close_date:
                where_clause += " AND s.escrowclosingdate <= :to_close_date"
                params["to_close_date"] = to_close_date

            if search:
                where_clause += """
                    AND (
                        LOWER(COALESCE(s.saleguid::text, '')) ILIKE :search
                        OR LOWER(
                            COALESCE(
                                CONCAT_WS(', ',
                                    CONCAT_WS(' ', sp.streetnumber, sp.streetaddress, sp.unit, sp.direction),
                                    sp.city,
                                    sp.state,
                                    sp.zip
                                ),
                                ''
                            )
                        ) ILIKE :search
                        OR LOWER(COALESCE(sp.state, '')) ILIKE :search
                        OR EXISTS (
                            SELECT 1
                            FROM sale_contact scb
                            WHERE scb.saleguid = s.saleguid
                              AND LOWER(COALESCE(scb.role, '')) = 'buyer'
                              AND LOWER(TRIM(COALESCE(scb.firstname, '') || ' ' || COALESCE(scb.lastname, ''))) ILIKE :search
                        )
                        OR EXISTS (
                            SELECT 1
                            FROM sale_contact scs
                            WHERE scs.saleguid = s.saleguid
                              AND LOWER(COALESCE(scs.role, '')) = 'seller'
                              AND LOWER(TRIM(COALESCE(scs.firstname, '') || ' ' || COALESCE(scs.lastname, ''))) ILIKE :search
                        )
                        OR EXISTS (
                            SELECT 1
                            FROM users r2
                            WHERE r2.userguid = s.reviewerguid
                              AND LOWER(TRIM(COALESCE(r2.firstname, '') || ' ' || COALESCE(r2.lastname, ''))) ILIKE :search
                        )
                    )
                """
                params["search"] = f"%{search.strip().lower()}%"

            count_query = f"""
                SELECT COUNT(DISTINCT s.saleguid) AS total
                FROM sale s
                LEFT JOIN sale_property sp ON sp.saleguid = s.saleguid
                {where_clause}
            """

            data_query = f"""
                SELECT
                    s.saleguid AS skyslopefileid,
                    s.saleprice AS ss_sale_price,
                    TRIM(s.status) AS ss_status,
                    s.escrowclosingdate AS ss_closed_date,
                    s.contractacceptancedate AS ss_contract_date,
                    s.listingprice AS ss_listing_price,
                    sp.state AS state,
                    CONCAT_WS(' ',
                        sp.streetnumber,
                        sp.streetaddress,
                        sp.unit,
                        sp.city,
                        sp.state,
                        sp.zip
                    ) AS property_address,
                    NULLIF(
                        TRIM(COALESCE(r.firstname, '') || ' ' || COALESCE(r.lastname, '')),
                        ''
                    ) AS reviewer
                FROM sale s
                LEFT JOIN users r ON s.reviewerguid = r.userguid
                LEFT JOIN sale_property sp ON sp.saleguid = s.saleguid
                {where_clause}
                ORDER BY s.saleguid
            """

            if page is not None and page_size is not None:
                offset = (page - 1) * page_size
                data_query += " LIMIT :limit OFFSET :offset"
                params["limit"] = page_size
                params["offset"] = offset

            count_params = {k: v for k, v in params.items() if k not in ("limit", "offset")}

            total = db.execute(text(count_query), count_params).scalar()
            rows = db.execute(text(data_query), params).mappings().all()
            rows = [dict(r) for r in rows]

            return {"mode": "skyslope_only", "total": total, "data": rows}

        base_cte = """
            WITH latest_reconciliation AS (
                SELECT DISTINCT ON (rd.transactionid)
                    rd.transactionid,
                    rd.be_source_table,
                    rd.saleguid,
                    rd.be_sale_price,
                    rd.skyslope_sale_price,
                    rd.sale_price_match,
                    rd.be_close_date_value,
                    rd.skyslope_close_date_value,
                    rd.close_date_match,
                    rd.be_contract_date,
                    rd.skyslope_contract_date,
                    rd.contract_date_match,
                    rd.be_listing_price,
                    rd.skyslope_listing_price,
                    rd.listing_price_match,
                    rd.be_status_value,
                    rd.skyslope_status_value,
                    rd.status_match,
                    rd.be_gross_commission,
                    rd.skyslope_gross_commission,
                    rd.gross_commission_match,
                    rd.be_buyer_name,
                    rd.skyslope_buyer_name,
                    rd.buyer_name_match,
                    rd.be_seller_name,
                    rd.skyslope_seller_name,
                    rd.seller_name_match,
                    rd.evaluated_at
                FROM reconciliation_data rd
                ORDER BY rd.transactionid, rd.evaluated_at DESC NULLS LAST
            ),
            brokerage_base AS (
                SELECT
                    'brokerage_engine'::text AS source_table,
                    be.transaction_identifier_transactionid AS transaction_id,
                    be.skyslopefileid::text AS skyslopefileid,
                    be.property_address,
                    be.state,
                    be.transaction_specialist,
                    be.transaction_status AS be_transaction_status,
                    be.tags
                FROM brokerage_engine be
            ),
            other_income_base AS (
                SELECT
                    'otherincome_transactions'::text AS source_table,
                    oit.transaction_identifier_transactionid AS transaction_id,
                    oit.skyslopefileid::text AS skyslopefileid,
                    oit.property_address,
                    oit.state,
                    oit.transaction_specialist,
                    oit.transaction_status AS be_transaction_status,
                    oit.tags
                FROM otherincome_transactions oit
            ),
            combined_source AS (
                SELECT * FROM brokerage_base
                UNION ALL
                SELECT * FROM other_income_base
            ),
            base AS (
                SELECT
                    cs.source_table,
                    cs.transaction_id,
                    cs.skyslopefileid,
                    COALESCE(
                        lr.saleguid,
                        CASE
                            WHEN cs.skyslopefileid ~* '^[0-9a-f-]{36}$'
                            THEN cs.skyslopefileid::uuid
                            ELSE NULL
                        END
                    ) AS saleguid,
                    cs.property_address,
                    cs.state,
                    cs.transaction_specialist,

                    lr.be_sale_price,
                    lr.skyslope_sale_price AS ss_sale_price,
                    lr.sale_price_match AS sale_price_comparison,

                    lr.be_close_date_value AS be_closed_date,
                    lr.skyslope_close_date_value AS ss_closed_date,
                    lr.close_date_match AS closed_date_comparison,

                    lr.be_contract_date,
                    lr.skyslope_contract_date AS ss_contract_date,
                    lr.contract_date_match AS contract_date_comparison,

                    lr.be_listing_price,
                    lr.skyslope_listing_price AS ss_listing_price,
                    lr.listing_price_match AS listing_price_comparison,

                    lr.be_status_value AS be_transaction_status,
                    lr.skyslope_status_value AS ss_transaction_status,
                    lr.status_match AS transaction_status_comparison,

                    lr.be_gross_commission,
                    lr.skyslope_gross_commission AS ss_gross_commission,
                    lr.gross_commission_match AS gross_commission_mismatch,

                    lr.be_buyer_name AS buyer_name,
                    lr.skyslope_buyer_name AS ss_buyer_name,
                    lr.buyer_name_match AS buyer_name_comparison,

                    lr.be_seller_name AS seller_name,
                    lr.skyslope_seller_name AS ss_seller_name,
                    lr.seller_name_match AS seller_name_comparison,

                    CASE
                        WHEN cs.tags ILIKE '%%titlepaymentreceived%%' THEN 'titlepaymentreceived'
                        WHEN cs.tags ILIKE '%%commissionverified%%' THEN 'commissionverified'
                        WHEN cs.tags ILIKE '%%cdasent%%' THEN 'cdasent'
                        WHEN cs.tags ILIKE '%%complete%%' THEN 'complete'
                        WHEN cs.tags ILIKE '%%open%%' THEN 'open'
                        ELSE NULL
                    END AS be_stage
                FROM combined_source cs
                LEFT JOIN latest_reconciliation lr
                    ON lr.transactionid = cs.transaction_id
            )
        """

        where_clause = " WHERE 1=1"
        params = {}

        if status != "all":
            where_clause += """
                AND (
                    CASE
                        WHEN b.be_transaction_status ILIKE 'pending'
                          OR b.be_transaction_status ILIKE 'active'
                          OR b.be_transaction_status ILIKE 'in_progress'
                        THEN 'pending'
                        WHEN b.be_transaction_status ILIKE 'closed'
                        THEN 'closed'
                        WHEN b.be_transaction_status ILIKE 'cancelled'
                          OR b.be_transaction_status ILIKE 'canceled'
                          OR b.be_transaction_status ILIKE 'canceled/app'
                          OR b.be_transaction_status ILIKE 'canceled/pend'
                        THEN 'cancelled'
                        ELSE 'other'
                    END = :status
                )
            """
            params["status"] = status

        if ps_list:
            placeholders = ", ".join(f":ps_{i}" for i in range(len(ps_list)))
            where_clause += f" AND b.be_stage IN ({placeholders})"
            for i, v in enumerate(ps_list):
                params[f"ps_{i}"] = v

        if state_list:
            placeholders = ", ".join(f":state_{i}" for i in range(len(state_list)))
            where_clause += f" AND LOWER(b.state) IN ({placeholders})"
            for i, v in enumerate(state_list):
                params[f"state_{i}"] = v.lower()

        if ts_list:
            unassigned_requested = any(v.lower() == "unassigned" for v in ts_list)
            named = [v for v in ts_list if v.lower() != "unassigned"]

            if unassigned_requested and named:
                placeholders = ", ".join(f":ts_{i}" for i in range(len(named)))
                where_clause += f"""
                    AND (
                        b.transaction_specialist IS NULL
                        OR b.transaction_specialist = ''
                        OR b.transaction_specialist IN ({placeholders})
                    )
                """
                for i, v in enumerate(named):
                    params[f"ts_{i}"] = v
            elif unassigned_requested:
                where_clause += """
                    AND (
                        b.transaction_specialist IS NULL
                        OR b.transaction_specialist = ''
                    )
                """
            else:
                placeholders = ", ".join(f":ts_{i}" for i in range(len(named)))
                where_clause += f" AND b.transaction_specialist IN ({placeholders})"
                for i, v in enumerate(named):
                    params[f"ts_{i}"] = v

        if from_close_date:
            where_clause += " AND b.be_closed_date >= :from_close_date"
            params["from_close_date"] = from_close_date

        if to_close_date:
            where_clause += " AND b.be_closed_date <= :to_close_date"
            params["to_close_date"] = to_close_date

        if search:
            where_clause += """
                AND (
                    COALESCE(b.transaction_id::text, '') ILIKE :search
                    OR COALESCE(b.property_address, '') ILIKE :search
                    OR COALESCE(b.state, '') ILIKE :search
                    OR COALESCE(b.transaction_specialist, '') ILIKE :search
                    OR COALESCE(b.buyer_name, '') ILIKE :search
                    OR COALESCE(b.seller_name, '') ILIKE :search
                    OR COALESCE(b.skyslopefileid::text, '') ILIKE :search
                    OR COALESCE(b.source_table, '') ILIKE :search
                )
            """
            params["search"] = f"%{search}%"

        count_query = base_cte + " SELECT COUNT(*) AS total FROM base b" + where_clause + ";"
        data_query = (
            base_cte
            + " SELECT * FROM base b"
            + where_clause
            + " ORDER BY b.transaction_id"
        )

        if page is not None and page_size is not None:
            offset = (page - 1) * page_size
            data_query += " LIMIT :limit OFFSET :offset;"
            params["limit"] = page_size
            params["offset"] = offset
        else:
            data_query += ";"

        count_params = {k: v for k, v in params.items() if k not in ("limit", "offset")}

        total = db.execute(text(count_query), count_params).scalar()
        rows = db.execute(text(data_query), params).mappings().all()
        rows = [dict(r) for r in rows]

        if mismatch:
            def has_mismatch(r):
                return any(
                    r.get(k) == "mismatch" for k in (
                        "sale_price_comparison",
                        "closed_date_comparison",
                        "contract_date_comparison",
                        "transaction_status_comparison",
                        "gross_commission_mismatch",
                        "buyer_name_comparison",
                        "seller_name_comparison",
                        "listing_price_comparison",
                    )
                )

            rows = [r for r in rows if has_mismatch(r)]
            total = len(rows)

        return {"mode": "full_comparison", "total": total, "data": rows}

    finally:
        pass


@router.get("/month-closing/listing")
def get_month_closing(
    status: str = "all",
    skyslope: bool = False,
    state: List[str] = Query(default=[]),
    from_close_date: str = None,
    to_close_date: str = None,
    transaction_specialist: List[str] = Query(default=[]),
    search: str = None,
    mismatch: bool = False,
    pending_subfilter: List[str] = Query(default=[]),
    page: int = 1,
    page_size: int = 50,
    db: Session = Depends(get_db)
):
    return fetch_month_closing_data(
        status=status,
        skyslope=skyslope,
        state=state,
        from_close_date=from_close_date,
        to_close_date=to_close_date,
        transaction_specialist=transaction_specialist,
        search=search,
        mismatch=mismatch,
        pending_subfilter=pending_subfilter,
        page=page,
        page_size=page_size,
        db=db
    )


@router.get("/month-closing/download")
def download_month_closing(
    status: str = "all",
    skyslope: bool = False,
    state: List[str] = Query(default=[]),
    from_close_date: str = None,
    to_close_date: str = None,
    transaction_specialist: List[str] = Query(default=[]),
    search: str = None,
    mismatch: bool = False,
    pending_subfilter: List[str] = Query(default=[]),
    db: Session = Depends(get_db)
):
    result = fetch_month_closing_data(
        status=status,
        skyslope=skyslope,
        state=state,
        from_close_date=from_close_date,
        to_close_date=to_close_date,
        transaction_specialist=transaction_specialist,
        search=search,
        mismatch=mismatch,
        pending_subfilter=pending_subfilter,
        page=None,
        page_size=None,
        db=db
    )

    data = result["data"]

    if skyslope:
        columns_map = {
            "skyslopefileid": "SkySlope File ID",
            "ss_sale_price": "SS Sale Price",
            "ss_status": "SS Status",
            "ss_closed_date": "SS Closed Date",
            "ss_contract_date": "SS Contract Date",
            "ss_listing_price": "SS Listing Price",
            "state": "State",
            "property_address": "Property Address",
            "reviewer": "Reviewer"
        }
    else:
        columns_map = {
            "transaction_id": "Transaction ID",
            "skyslopefileid": "SkySlope File ID",
            "property_address": "Property Address",
            "state": "State",
            "transaction_specialist": "Transaction Specialist",
            "be_stage": "BE Stage",
            "be_sale_price": "BE Sale Price",
            "ss_sale_price": "SS Sale Price",
            "sale_price_comparison": "Sale Price Comparison",
            "be_closed_date": "BE Closed Date",
            "ss_closed_date": "SS Closed Date",
            "closed_date_comparison": "Closed Date Comparison",
            "be_contract_date": "BE Contract Date",
            "ss_contract_date": "SS Contract Date",
            "contract_date_comparison": "Contract Date Comparison",
            "be_listing_price": "BE Listing Price",
            "ss_listing_price": "SS Listing Price",
            "listing_price_comparison": "Listing Price Comparison",
            "be_transaction_status": "BE Transaction Status",
            "ss_transaction_status": "SS Transaction Status",
            "transaction_status_comparison": "Transaction Status Comparison",
            "be_gross_commission": "BE Gross Commission",
            "ss_gross_commission": "SS Gross Commission",
            "gross_commission_mismatch": "Gross Commission Mismatch",
            "buyer_name": "BE Buyer Name",
            "ss_buyer_name": "SS Buyer Name",
            "buyer_name_comparison": "Buyer Name Comparison",
            "seller_name": "BE Seller Name",
            "ss_seller_name": "SS Seller Name",
            "seller_name_comparison": "Seller Name Comparison"
        }

    rows_to_export = []
    for r in data:
        row_dict = {}
        for key, header in columns_map.items():
            val = r.get(key)
            if isinstance(val, Decimal):
                val = float(val)
            elif isinstance(val, (datetime.date, datetime.datetime)):
                val = val.strftime("%Y-%m-%d")
            elif isinstance(val, bool):
                val = "Yes" if val else "No"
            elif val is None:
                val = ""
            row_dict[header] = val
        rows_to_export.append(row_dict)

    df = pd.DataFrame(rows_to_export)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Month Closing Report", index=False)

        worksheet = writer.sheets["Month Closing Report"]
        worksheet.views.sheetView[0].showGridLines = True

        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

        font_body = Font(name="Segoe UI", size=10)
        fill_even = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        fill_mismatch = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
        font_mismatch = Font(name="Segoe UI", size=10, bold=True, color="C53929")

        thin_border = Border(
            left=Side(style='thin', color='D0D5DD'),
            right=Side(style='thin', color='D0D5DD'),
            top=Side(style='thin', color='D0D5DD'),
            bottom=Side(style='thin', color='D0D5DD')
        )

        worksheet.row_dimensions[1].height = 28
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_header
            cell.border = thin_border

        currency_cols = []
        date_cols = []
        center_cols = []

        currency_keywords = ["gross commission", "sale price", "listing price"]
        date_keywords = ["closed date", "contract date"]
        center_keywords = ["id", "comparison", "mismatch", "state", "status", "stage"]

        for idx, col_name in enumerate(df.columns):
            col_name_lower = col_name.lower()
            if any(kw in col_name_lower for kw in currency_keywords):
                currency_cols.append(idx + 1)
            elif any(kw in col_name_lower for kw in date_keywords):
                date_cols.append(idx + 1)
            elif any(kw in col_name_lower for kw in center_keywords):
                center_cols.append(idx + 1)

        for row_num in range(2, len(df) + 2):
            worksheet.row_dimensions[row_num].height = 20
            is_even_row = (row_num % 2 == 0)

            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = font_body
                cell.border = thin_border

                if is_even_row:
                    cell.fill = fill_even

                val = cell.value
                val_str = str(val).strip().lower() if val is not None else ""
                col_name = df.columns[col_num - 1]
                col_name_lower = col_name.lower()

                is_cell_mismatch = False
                if any(kw in col_name_lower for kw in ["comparison", "mismatch"]):
                    if val_str in ["yes", "mismatch"]:
                        is_cell_mismatch = True

                if is_cell_mismatch:
                    cell.fill = fill_mismatch
                    cell.font = font_mismatch

                if col_num in currency_cols:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(val, (int, float)):
                        cell.number_format = '$#,##0.00'
                elif col_num in date_cols:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_num in center_cols:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = cell.value
                if val is not None:
                    if cell.column in currency_cols and isinstance(val, (int, float)):
                        val_len = len(f"${val:,.2f}")
                    else:
                        val_len = len(str(val))
                    if val_len > max_len:
                        max_len = val_len
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output.seek(0)

    filename = f"month_closing_report_{status}.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )