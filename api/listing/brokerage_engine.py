from fastapi import HTTPException, APIRouter, Query, Depends, Response
from db import get_db
from services.loaders import get_be_sync
import pandas as pd
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal
import datetime

router = APIRouter()

@router.get("/brokerage_engine/sync_info")
def brokerage_engine_sync_info(conn=Depends(get_db)):
    return get_be_sync(conn)

@router.get("/brokerage_engine")
def brokerage_engine(
    brokerhold: bool = Query(default=False),
    page: int = Query(default=1, ge=1),
    from_close_date: str = Query(default=None),
    to_close_date: str = Query(default=None),
    from_contract_date: str = Query(default=None),
    to_contract_date: str = Query(default=None),
    status: str = Query(default=None),
    search: str = Query(default=None),
    conn=Depends(get_db)
):
    cursor = conn.cursor()

    limit = 50
    offset = (page - 1) * limit

    base_query = """
        FROM brokerage_engine
        WHERE 1=1
    """

    params = []

    if brokerhold:
        base_query += " AND LOWER(tags) LIKE %s"
        params.append("%brokerhold%")

    if status:
        base_query += " AND LOWER(transaction_status) = %s"
        params.append(status.lower())

    if from_close_date:
        base_query += " AND closed_date >= %s"
        params.append(from_close_date)

    if to_close_date:
        base_query += " AND closed_date <= %s"
        params.append(to_close_date)

    if from_contract_date:
        base_query += " AND contract_date >= %s"
        params.append(from_contract_date)

    if to_contract_date:
        base_query += " AND contract_date <= %s"
        params.append(to_contract_date)

    if search:
        search_value = f"%{search.lower()}%"
        base_query += """
            AND (
                LOWER(transaction_identifier_transactionid::text) LIKE %s
                OR LOWER(COALESCE(property_address, '')) LIKE %s
                OR LOWER(COALESCE(buying_agent_name, '')) LIKE %s
            )
        """
        params.extend([search_value, search_value, search_value])

    count_query = "SELECT COUNT(*) " + base_query
    cursor.execute(count_query, params)
    total_count = cursor.fetchone()[0]

    data_query = """
        SELECT
            transaction_identifier_transactionid AS transactionid,
            property_address,
            buying_agent_name,
            sale_price,
            contract_date,
            closed_date AS close_date,
            transaction_specialist,
            transaction_status AS status,
            skyslopefileid
    """ + base_query

    data_query += " ORDER BY closed_date DESC NULLS LAST"
    data_query += " LIMIT %s OFFSET %s"

    data_params = params + [limit, offset]

    cursor.execute(data_query, data_params)

    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()
    data = [dict(zip(columns, row)) for row in rows]

    status_query = """
        SELECT DISTINCT transaction_status AS status
        FROM brokerage_engine
        WHERE transaction_status IS NOT NULL
        ORDER BY transaction_status
    """
    cursor.execute(status_query)
    status_list = [row[0] for row in cursor.fetchall()]

    return {
        "total_count": total_count,
        "filters": {
            "status_list": status_list
        },
        "data": data
    }

def norm(x):
    return str(x or "").replace("\u00A0", "").strip().lower()

@router.get("/brokerage_engine/detail")
def brokerage_detail(transactionid: str, conn=Depends(get_db)):
    cursor = conn.cursor()

    query = """
        SELECT
            be.skyslopefileid,
            be.listingguid,
            be.listing_office,
            be.transaction_identifier_transactionid,
            be.sale_price,
            be.closed_date,
            be.listing_price,
            be.contract_date,
            be.tags,
            be.buyer_name,
            be.seller_name,
            be.buying_agent_name,
            be.total_gross_commission,
            be.transaction_specialist,
            be.property_address,
            be.da_title_company,
            be.transaction_status,

            s.saleguid,
            s.listingprice,
            s.saleprice,
            s.mlsnumber,
            s.status,
            s.contractacceptancedate,
            s.escrowclosingdate,
            s.reviewerguid,
            s.agentguid,

            COALESCE(r.firstname || ' ' || r.lastname, '') AS reviewer_full_name,

            COALESCE(
                (
                    SELECT TRIM(COALESCE(uu.firstname, '') || ' ' || COALESCE(uu.lastname, ''))
                    FROM users uu
                    WHERE uu.userguid = s.agentguid
                ),
                ''
            ) AS skyslope_buying_agent_name,

            CONCAT_WS(', ',
                CONCAT_WS(' ', sp.streetnumber, sp.streetaddress),
                sp.city,
                sp.state,
                sp.zip
            ) AS propertyaddress,

            COALESCE(scn.officeGrossCommissionOnSale, 0) AS officegrosscommissiononsale

        FROM brokerage_engine be

        LEFT JOIN sale s
            ON be.skyslopefileid = s.saleguid

        LEFT JOIN users r
            ON s.reviewerguid = r.userguid

        LEFT JOIN sale_property sp
            ON s.saleguid = sp.saleguid

        LEFT JOIN sale_commission scn
            ON scn.saleguid = s.saleguid

        WHERE be.transaction_identifier_transactionid = %s
    """

    cursor.execute(query, (transactionid,))
    row = cursor.fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Transaction not found")

    columns = [desc[0] for desc in cursor.description]
    data = dict(zip(columns, row))

    skyslope_match = data.get("saleguid") is not None

    return {
        "transactionid": transactionid,
        "brokerage_engine": {
            "property_address": data.get("property_address"),
            "sale_price": data.get("sale_price"),
            "listing_price": data.get("listing_price"),
            "office": data.get("listing_office"),
            "buyer": data.get("buyer_name"),
            "seller": data.get("seller_name"),
            "buying_agent_name": data.get("buying_agent_name"),
            "contract_date": data.get("contract_date"),
            "closed_date": data.get("closed_date"),
            "tags": data.get("tags"),
            "status": data.get("transaction_status"),
            "transaction_specialist": data.get("transaction_specialist"),
            "skyslopefileid": data.get("skyslopefileid")    
        },
        "skyslope": {
            "match": skyslope_match,
            "saleguid": data.get("saleguid"),
            "property_address": data.get("propertyaddress"),
            "listingprice": data.get("listingprice"),
            "saleprice": data.get("saleprice"),
            "mlsnumber": data.get("mlsnumber"),
            "seller": data.get("seller_name"),
            "buyer": data.get("buyer_name"),
            "buying_agent": data.get("skyslope_buying_agent_name"),
            "reviewer_full_name": data.get("reviewer_full_name"),
            "status": data.get("status"),
            "contractacceptancedate": data.get("contractacceptancedate"),
            "escrowclosingdate": data.get("escrowclosingdate"),
            "canceldate": data.get("canceldate"),
            "officegrosscommissiononsale": data.get("officegrosscommissiononsale")
        }
    }

@router.get("/sale/noskyslopefileid/download")
def download_sale_no_skyslopefileid(conn=Depends(get_db)):
    cursor = conn.cursor()

    query = """
        SELECT *
        FROM brokerage_engine
        WHERE skyslopefileid IS NULL
        ORDER BY closed_date DESC NULLS LAST, contract_date DESC NULLS LAST
    """

    cursor.execute(query)
    data = cursor.fetchall()

    columns = [desc[0] for desc in cursor.description]

    rows_to_export = []
    for row in data:
        row_dict = dict(zip(columns, row))
        export_row = {}

        for col in columns:
            val = row_dict.get(col)

            if isinstance(val, Decimal):
                val = float(val)
            elif isinstance(val, (datetime.date, datetime.datetime)):
                val = val.strftime("%Y-%m-%d")
            elif isinstance(val, bool):
                val = "Yes" if val else "No"
            elif val is None:
                val = ""

            export_row[col] = val

        rows_to_export.append(export_row)

    df = pd.DataFrame(rows_to_export, columns=columns)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="No SkySlope File ID", index=False)

        worksheet = writer.sheets["No SkySlope File ID"]

        worksheet.views.sheetView[0].showGridLines = True
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

        font_body = Font(name="Segoe UI", size=10)
        fill_even = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin", color="D0D5DD"),
            right=Side(style="thin", color="D0D5DD"),
            top=Side(style="thin", color="D0D5DD"),
            bottom=Side(style="thin", color="D0D5DD")
        )

        worksheet.row_dimensions[1].height = 28
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_header
            cell.border = thin_border

        currency_cols = []
        date_cols = []
        center_cols = []

        currency_keywords = ["price", "commission", "amount", "gross", "net"]
        date_keywords = ["date"]
        center_keywords = ["id", "guid", "status"]

        for idx, col_name in enumerate(df.columns):
            col_name_lower = col_name.lower()
            if any(kw in col_name_lower for kw in currency_keywords):
                currency_cols.append(idx + 1)
            elif any(kw in col_name_lower for kw in date_keywords):
                date_cols.append(idx + 1)
            elif any(kw in col_name_lower for kw in center_keywords):
                center_cols.append(idx + 1)

        for row_num in range(2, len(df) + 2):
            worksheet.row_dimensions[row_num].height = 20
            is_even_row = (row_num % 2 == 0)

            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = font_body
                cell.border = thin_border

                if is_even_row:
                    cell.fill = fill_even

                val = cell.value

                if col_num in currency_cols:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(val, (int, float)):
                        cell.number_format = '$#,##0.00'
                elif col_num in date_cols:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_num in center_cols:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                val = cell.value
                if val is not None:
                    if cell.column in currency_cols and isinstance(val, (int, float)):
                        val_len = len(f"${val:,.2f}")
                    else:
                        val_len = len(str(val))
                    if val_len > max_len:
                        max_len = val_len

            worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    output.seek(0)

    filename = "sale_no_skyslopefileid_report.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )