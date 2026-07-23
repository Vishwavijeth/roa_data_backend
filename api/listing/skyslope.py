from fastapi import APIRouter, HTTPException, Query, Depends, Response
from db import get_db
from services.loaders import get_skyslope_sync
import io
import datetime
from decimal import Decimal
import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


router = APIRouter()


def norm(x):
    return str(x or "").replace("\u00A0", "").strip().lower()


def apply_skyslope_filters(
    base_filter: str,
    params: list,
    from_close_date=None,
    to_close_date=None,
    from_contract_date=None,
    to_contract_date=None,
    status=None,
    search=None,
    not_in_be: bool = False,
):
    if status:
        base_filter += " AND LOWER(s.status) = %s"
        params.append(status.lower())

    if from_close_date:
        base_filter += " AND s.escrowclosingdate >= %s"
        params.append(from_close_date)

    if to_close_date:
        base_filter += " AND s.escrowclosingdate <= %s"
        params.append(to_close_date)

    if from_contract_date:
        base_filter += " AND s.contractacceptancedate >= %s"
        params.append(from_contract_date)

    if to_contract_date:
        base_filter += " AND s.contractacceptancedate <= %s"
        params.append(to_contract_date)

    if search:
        search_value = f"%{search.lower()}%"
        base_filter += """
            AND (
                EXISTS (
                    SELECT 1
                    FROM brokerage_engine be
                    WHERE be.skyslopefileid = s.saleguid
                      AND LOWER(be.transaction_identifier_transactionid::text) LIKE %s
                )
                OR EXISTS (
                    SELECT 1
                    FROM sale_property sp
                    WHERE sp.saleguid = s.saleguid
                      AND LOWER(
                        CONCAT_WS(', ',
                            CONCAT_WS(' ', sp.streetnumber, sp.streetaddress),
                            sp.city,
                            sp.state,
                            sp.zip
                        )
                      ) LIKE %s
                )
                OR LOWER(s.saleguid::text) LIKE %s
            )
        """
        params.extend([search_value, search_value, search_value])

    if not_in_be:
        base_filter += """
            AND NOT EXISTS (
                SELECT 1
                FROM brokerage_engine be2
                WHERE be2.skyslopefileid = s.saleguid
            )
            AND NOT EXISTS (
                SELECT 1
                FROM otherincome_transactions oit
                WHERE oit.skyslopefileid = s.saleguid
            )
        """

    return base_filter, params


@router.get("/skyslope/sync_info")
def skyslope_sync_info(conn=Depends(get_db)):
    sync_info = get_skyslope_sync(conn)

    return {
        "sync_info": sync_info,
    }


@router.get("/skyslope-listing")
def skyslope_api(
    page: int = Query(default=1, ge=1),
    from_close_date: str = Query(default=None),
    to_close_date: str = Query(default=None),
    from_contract_date: str = Query(default=None),
    to_contract_date: str = Query(default=None),
    status: str = Query(default=None),
    search: str = Query(default=None),
    not_in_be: bool = Query(default=False),
    conn=Depends(get_db)
):
    cursor = conn.cursor()

    limit = 50
    offset = (page - 1) * limit

    base_filter = """
        FROM sale s
        WHERE 1=1
    """

    params = []

    base_filter, params = apply_skyslope_filters(
        base_filter=base_filter,
        params=params,
        from_close_date=from_close_date,
        to_close_date=to_close_date,
        from_contract_date=from_contract_date,
        to_contract_date=to_contract_date,
        status=status,
        search=search,
        not_in_be=not_in_be,
    )

    count_query = "SELECT COUNT(*) " + base_filter
    cursor.execute(count_query, params)
    total_count = cursor.fetchone()[0]

    data_query = """
        SELECT
            s.saleguid,
            CONCAT_WS(', ',
                CONCAT_WS(' ', sp.streetnumber, sp.streetaddress),
                sp.city,
                sp.state,
                sp.zip
            ) AS propertyaddress,
            s.escrowclosingdate AS close_date,
            s.status,
            CASE
                WHEN be.transaction_identifier_transactionid IS NULL
                THEN 'No related BE data'
                ELSE be.transaction_identifier_transactionid::text
            END AS transaction_id,
            NULLIF(
                COALESCE(
                    (
                        SELECT STRING_AGG(
                            TRIM(
                                COALESCE(sc.firstname, '') || ' ' ||
                                COALESCE(sc.lastname, '')
                            ),
                            ', '
                        )
                        FROM sale_contact sc
                        WHERE sc.saleguid = s.saleguid
                          AND LOWER(sc.role) = 'buyer'
                    ),
                    ''
                ),
                ''
            ) AS buyer_name,
            NULLIF(
                (
                    SELECT TRIM(COALESCE(u.firstname, '') || ' ' || COALESCE(u.lastname, ''))
                    FROM users u
                    WHERE u.userguid = s.agentguid
                ),
                ''
            ) AS buyer_agent_name,
            s.contractacceptancedate AS contract_date,
            NULLIF(
                TRIM(COALESCE(r.firstname, '') || ' ' || COALESCE(r.lastname, '')),
                ''
            ) AS reviewer
        FROM sale s
        LEFT JOIN users u ON s.agentguid = u.userguid
        LEFT JOIN users r ON s.reviewerguid = r.userguid
        LEFT JOIN sale_property sp ON s.saleguid = sp.saleguid
        LEFT JOIN brokerage_engine be ON be.skyslopefileid = s.saleguid
    """

    data_query += base_filter.replace("FROM sale s", "")

    data_query += """
        ORDER BY s.escrowclosingdate DESC NULLS LAST
        LIMIT %s OFFSET %s
    """

    cursor.execute(data_query, params + [limit, offset])

    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()
    data = [dict(zip(columns, row)) for row in rows]

    status_query = """
        SELECT DISTINCT status
        FROM sale
        WHERE status IS NOT NULL
        ORDER BY status
    """
    cursor.execute(status_query)
    status_list = [row[0] for row in cursor.fetchall()]

    return {
        "total_count": total_count,
        "filters": {
            "status_list": status_list,
            "not_in_be": not_in_be,
        },
        "data": data
    }


@router.get("/skyslope/detail")
def skyslope_detail(saleguid: str, conn=Depends(get_db)):
    cursor = conn.cursor()

    skyslope_query = """
        SELECT
            s.saleguid,
            s.saleguid AS skyslope_saleguid,
            CONCAT_WS(', ',
                CONCAT_WS(' ', sp.streetnumber, sp.streetaddress),
                sp.city,
                sp.state,
                sp.zip
            ) AS propertyaddress,
            s.listingprice,
            s.saleprice,
            s.mlsnumber,
            s.contractacceptancedate,
            s.escrowclosingdate,
            s.status,

            COALESCE(u.firstname || ' ' || u.lastname, NULL) AS buyer_agent_name,
            COALESCE(u.email, NULL) AS buyer_agent_email,
            COALESCE(r.firstname || ' ' || r.lastname, NULL) AS reviewer,

            COALESCE(
                (
                    SELECT STRING_AGG(
                        TRIM(
                            COALESCE(sc.firstname, '') || ' ' ||
                            COALESCE(sc.lastname, '')
                        ),
                        ', '
                    )
                    FROM sale_contact sc
                    WHERE sc.saleguid = s.saleguid
                      AND LOWER(sc.role) = 'buyer'
                ),
                NULL
            ) AS buyer,

            COALESCE(
                (
                    SELECT STRING_AGG(
                        TRIM(
                            COALESCE(sc.firstname, '') || ' ' ||
                            COALESCE(sc.lastname, '')
                        ),
                        ', '
                    )
                    FROM sale_contact sc
                    WHERE sc.saleguid = s.saleguid
                      AND LOWER(sc.role) = 'seller'
                ),
                NULL
            ) AS seller
        FROM sale s
        LEFT JOIN users u ON s.agentguid = u.userguid
        LEFT JOIN users r ON s.reviewerguid = r.userguid
        LEFT JOIN sale_property sp ON s.saleguid = sp.saleguid
        WHERE s.saleguid = %s
    """

    cursor.execute(skyslope_query, [saleguid])
    row = cursor.fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Sale not found")

    columns = [desc[0] for desc in cursor.description]
    skyslope_data = dict(zip(columns, row))

    be_query = """
        SELECT
            be.transaction_identifier_transactionid::text AS transactionid,
            be.property_address,
            be.sale_price,
            be.listing_price,
            be.listing_office AS office,
            be.buyer_name,
            be.seller_name,
            be.buying_agent_name,
            be.contract_date,
            be.closed_date,
            be.tags,
            be.transaction_status,
            be.transaction_specialist,
            be.skyslopefileid,
            be.total_gross_commission,
            CASE
                WHEN LOWER(be.transaction_status) = 'cancelled'
                    THEN 'ignored_cancelled_duplicate'
                ELSE 'active'
            END AS record_role
        FROM brokerage_engine be
        WHERE be.skyslopefileid = %s
        ORDER BY
            CASE
                WHEN LOWER(be.transaction_status) = 'cancelled' THEN 2
                ELSE 1
            END,
            be.closed_date DESC NULLS LAST
    """

    cursor.execute(be_query, [saleguid])
    be_rows = cursor.fetchall()
    be_columns = [desc[0] for desc in cursor.description]
    brokerage_engine_records = [dict(zip(be_columns, r)) for r in be_rows]

    other_income_query = """
        SELECT
            oit.transaction_identifier_transactionid::text AS transactionid,
            oit.property_address,
            oit.income_type,
            oit.income_received_date,
            oit.income_received,
            oit.gross_commission,
            oit.agent_net,
            oit.brokerage_net,
            oit.agents,
            oit.client_type,
            oit.client_name,
            oit.client_phone,
            oit.client_email,
            oit.tags,
            oit.effective_at,
            oit.finalized_date,
            oit.transaction_specialist,
            oit.transaction_status,
            oit.skyslopefileid,
            'other_income' AS record_role
        FROM otherincome_transactions oit
        WHERE oit.skyslopefileid = %s
        ORDER BY oit.finalized_date DESC NULLS LAST, oit.income_received_date DESC NULLS LAST
    """

    cursor.execute(other_income_query, [saleguid])
    oi_rows = cursor.fetchall()
    oi_columns = [desc[0] for desc in cursor.description]
    otherincome_records = [dict(zip(oi_columns, r)) for r in oi_rows]

    return {
        "saleguid": skyslope_data["saleguid"],
        "skyslope": {
            "saleguid": skyslope_data["skyslope_saleguid"],
            "propertyaddress": skyslope_data["propertyaddress"],
            "listingprice": skyslope_data["listingprice"],
            "saleprice": skyslope_data["saleprice"],
            "mlsnumber": skyslope_data["mlsnumber"],
            "seller": skyslope_data["seller"],
            "buyer": skyslope_data["buyer"],
            "buyer_agent": skyslope_data["buyer_agent_name"],
            "buyer_agent_email": skyslope_data["buyer_agent_email"],
            "reviewer": skyslope_data["reviewer"],
            "status": skyslope_data["status"],
            "contractacceptancedate": skyslope_data["contractacceptancedate"],
            "escrowclosingdate": skyslope_data["escrowclosingdate"]
        },
        "brokerage_engine_records": brokerage_engine_records,
        "otherincome_transactions": otherincome_records,
    }


@router.get("/skyslope/download")
def skyslope_download(
    from_close_date: str = Query(default=None),
    to_close_date: str = Query(default=None),
    from_contract_date: str = Query(default=None),
    to_contract_date: str = Query(default=None),
    status: str = Query(default=None),
    search: str = Query(default=None),
    not_in_be: bool = Query(default=False),
    conn=Depends(get_db)
):
    cursor = conn.cursor()

    base_filter = """
        FROM sale s
        WHERE 1=1
    """

    params = []

    base_filter, params = apply_skyslope_filters(
        base_filter=base_filter,
        params=params,
        from_close_date=from_close_date,
        to_close_date=to_close_date,
        from_contract_date=from_contract_date,
        to_contract_date=to_contract_date,
        status=status,
        search=search,
        not_in_be=not_in_be,
    )

    data_query = """
        SELECT
            s.saleguid AS saleguid,
            CONCAT_WS(', ',
                CONCAT_WS(' ', sp.streetnumber, sp.streetaddress),
                sp.city,
                sp.state,
                sp.zip
            ) AS property_address,

            NULLIF(
                (
                    SELECT STRING_AGG(
                        TRIM(COALESCE(sc.firstname, '') || ' ' || COALESCE(sc.lastname, '')),
                        ', '
                        ORDER BY sc.firstname, sc.lastname
                    )
                    FROM sale_contact sc
                    WHERE sc.saleguid = s.saleguid
                      AND LOWER(sc.role) = 'buyer'
                ),
                ''
            ) AS buyer_name,

            NULLIF(
                (
                    SELECT STRING_AGG(
                        TRIM(COALESCE(sc.firstname, '') || ' ' || COALESCE(sc.lastname, '')),
                        ', '
                        ORDER BY sc.firstname, sc.lastname
                    )
                    FROM sale_contact sc
                    WHERE sc.saleguid = s.saleguid
                      AND LOWER(sc.role) = 'seller'
                ),
                ''
            ) AS seller_name,

            s.saleprice AS sale_price,
            s.listingprice AS listing_price,
            s.escrowclosingdate AS escrow_close_date,
            s.contractacceptancedate AS contract_date,
            s.status AS status,
            st.name AS stage,
            s.dealtype AS dealtype,

            sc2."officegrosscommissiononsale" AS office_gross_commission_on_sale,
            sc2."salecommissionamount" AS sale_commission_amount,
            sc2."listingcommissionamount" AS listing_commission_amount
        FROM sale s
        LEFT JOIN sale_property sp ON s.saleguid = sp.saleguid
        LEFT JOIN stage st ON s.stageid = st.stageid
        LEFT JOIN sale_commission sc2 ON sc2.saleguid = s.saleguid
    """

    data_query += base_filter.replace("FROM sale s", "")

    data_query += """
        ORDER BY s.escrowclosingdate DESC NULLS LAST, s.saleguid
    """

    cursor.execute(data_query, params)
    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()
    data = [dict(zip(columns, row)) for row in rows]

    columns_map = {
        "saleguid": "Sale GUID",
        "property_address": "Property Address",
        "buyer_name": "Buyer Name",
        "seller_name": "Seller Name",
        "sale_price": "Sale Price",
        "listing_price": "Listing Price",
        "escrow_close_date": "Escrow Close Date",
        "contract_date": "Contract Date",
        "status": "Status",
        "stage": "Stage",
        "dealtype": "Deal Type",
        "office_gross_commission_on_sale": "Office Gross Commission On Sale",
        "sale_commission_amount": "Sale Commission Amount",
        "listing_commission_amount": "Listing Commission Amount",
    }

    rows_to_export = []
    for record in data:
        row_dict = {}
        for key, header in columns_map.items():
            val = record.get(key)

            if isinstance(val, Decimal):
                val = float(val)
            elif isinstance(val, (datetime.date, datetime.datetime)):
                val = val.strftime("%Y-%m-%d")
            elif isinstance(val, bool):
                val = "Yes" if val else "No"
            elif val is None:
                val = ""

            row_dict[header] = val

        rows_to_export.append(row_dict)

    df = pd.DataFrame(rows_to_export)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="SkySlope Report", index=False)

        worksheet = writer.sheets["SkySlope Report"]
        worksheet.freeze_panes = "A2"

        font_header = Font(name="Segoe UI", size=11, bold=True)
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        font_body = Font(name="Segoe UI", size=10)

        worksheet.row_dimensions[1].height = 28

        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = font_header
            cell.alignment = align_header

        currency_cols = []
        date_cols = []
        center_cols = []

        currency_keywords = [
            "sale price",
            "listing price",
            "commission",
        ]
        date_keywords = ["date"]
        center_keywords = ["sale guid", "status", "stage", "deal type"]

        for idx, col_name in enumerate(df.columns):
            col_name_lower = col_name.lower()
            if any(kw in col_name_lower for kw in currency_keywords):
                currency_cols.append(idx + 1)
            elif any(kw in col_name_lower for kw in date_keywords):
                date_cols.append(idx + 1)
            elif any(kw in col_name_lower for kw in center_keywords):
                center_cols.append(idx + 1)

        for row_num in range(2, len(df) + 2):
            worksheet.row_dimensions[row_num].height = 20

            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = font_body
                val = cell.value

                if col_num in currency_cols:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(val, (int, float)):
                        cell.number_format = '$#,##0.00'
                elif col_num in date_cols:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_num in center_cols:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                val = cell.value
                if val is not None:
                    if cell.column in currency_cols and isinstance(val, (int, float)):
                        val_len = len(f"${val:,.2f}")
                    else:
                        val_len = len(str(val))
                    if val_len > max_len:
                        max_len = val_len

            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output.seek(0)

    filename = "skyslope_sale_report.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )