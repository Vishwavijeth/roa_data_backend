from typing import List, Optional
from fastapi import Query, HTTPException, APIRouter, Depends, Response
from db import get_db
import pandas as pd
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal
import datetime

router = APIRouter()

@router.get("/otherincome_transactions")
def otherincome_transactions(
    page: int = Query(default=1, ge=1),
    from_income_received_date: str = Query(default=None),
    to_income_received_date: str = Query(default=None),
    from_finalized_date: str = Query(default=None),
    to_finalized_date: str = Query(default=None),
    status: Optional[List[str]] = Query(default=None),
    income_type: Optional[List[str]] = Query(default=None),
    search: str = Query(default=None),
    conn=Depends(get_db)
):
    cursor = conn.cursor()

    limit = 50
    offset = (page - 1) * limit

    base_query = """
        FROM otherincome_transactions
        WHERE 1=1
    """

    params = []

    if status:
        status_placeholders = ", ".join(["%s"] * len(status))
        base_query += f" AND LOWER(transaction_status) IN ({status_placeholders})"
        params.extend([s.lower() for s in status])

    if income_type:
        income_type_placeholders = ", ".join(["%s"] * len(income_type))
        base_query += f" AND LOWER(income_type) IN ({income_type_placeholders})"
        params.extend([i.lower() for i in income_type])

    if from_income_received_date:
        base_query += " AND income_received_date >= %s"
        params.append(from_income_received_date)

    if to_income_received_date:
        base_query += " AND income_received_date <= %s"
        params.append(to_income_received_date)

    if from_finalized_date:
        base_query += " AND finalized_date >= %s"
        params.append(from_finalized_date)

    if to_finalized_date:
        base_query += " AND finalized_date <= %s"
        params.append(to_finalized_date)

    if search:
        search_value = f"%{search.lower()}%"
        base_query += """
            AND (
                LOWER(transaction_identifier_transactionid::text) LIKE %s
                OR LOWER(COALESCE(property_address, '')) LIKE %s
                OR LOWER(COALESCE(client_name, '')) LIKE %s
                OR LOWER(COALESCE(client_email, '')) LIKE %s
                OR LOWER(COALESCE(agents, '')) LIKE %s
                OR LOWER(COALESCE(transaction_specialist, '')) LIKE %s
                OR LOWER(COALESCE(income_type, '')) LIKE %s
            )
        """
        params.extend([
            search_value, search_value, search_value, search_value,
            search_value, search_value, search_value
        ])

    count_query = "SELECT COUNT(*) " + base_query
    cursor.execute(count_query, params)
    total_count = cursor.fetchone()[0]

    data_query = """
        SELECT
            transaction_identifier_transactionid AS transactionid,
            property_address,
            income_type,
            income_received_date,
            income_received,
            gross_commission,
            agent_net,
            brokerage_net,
            client_name,
            agents,
            finalized_date,
            transaction_specialist,
            transaction_status AS status,
            skyslopefileid
    """ + base_query

    data_query += " ORDER BY finalized_date DESC NULLS LAST, income_received_date DESC NULLS LAST"
    data_query += " LIMIT %s OFFSET %s"

    data_params = params + [limit, offset]

    cursor.execute(data_query, data_params)

    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()
    data = [dict(zip(columns, row)) for row in rows]

    filter_query = """
        SELECT DISTINCT transaction_status
        FROM otherincome_transactions
        WHERE transaction_status IS NOT NULL
        ORDER BY transaction_status
    """
    cursor.execute(filter_query)
    status_options = [row[0] for row in cursor.fetchall()]

    income_type_query = """
        SELECT DISTINCT income_type
        FROM otherincome_transactions
        WHERE income_type IS NOT NULL
        ORDER BY income_type
    """
    cursor.execute(income_type_query)
    income_type_options = [row[0] for row in cursor.fetchall()]

    return {
        "total_count": total_count,
        "filters": {
            "status": status_options,
            "income_type": income_type_options
        },
        "data": data
    }


@router.get("/otherincome_transactions/detail")
def otherincome_detail(transactionid: str, conn=Depends(get_db)):
    cursor = conn.cursor()

    query = """
        SELECT
            oit.skyslopefileid,
            oit.listingguid,
            oit.transaction_identifier_transactionid,
            oit.property_address,
            oit.income_type,
            oit.income_received_date,
            oit.income_received,
            oit.gross_commission,
            oit.agent_net,
            oit.brokerage_net,
            oit.agents,
            oit.tags,
            oit.effective_at,
            oit.finalized_date,
            oit.transaction_specialist,
            oit.transaction_status,

            s.saleguid,
            s.listingprice,
            s.saleprice,
            s.mlsnumber,
            s.status,
            s.contractacceptancedate,
            s.escrowclosingdate,
            s.reviewerguid,
            s.agentguid,

            COALESCE(r.firstname || ' ' || r.lastname, '') AS reviewer_full_name,

            COALESCE(
                (
                    SELECT TRIM(COALESCE(uu.firstname, '') || ' ' || COALESCE(uu.lastname, ''))
                    FROM users uu
                    WHERE uu.userguid = s.agentguid
                ),
                ''
            ) AS skyslope_buying_agent_name,

            CONCAT_WS(', ',
                CONCAT_WS(' ', sp.streetnumber, sp.streetaddress),
                sp.city,
                sp.state,
                sp.zip
            ) AS propertyaddress,

            COALESCE(scn.officeGrossCommissionOnSale, 0) AS officegrosscommissiononsale

        FROM otherincome_transactions oit

        LEFT JOIN sale s
            ON oit.skyslopefileid = s.saleguid

        LEFT JOIN users r
            ON s.reviewerguid = r.userguid

        LEFT JOIN sale_property sp
            ON s.saleguid = sp.saleguid

        LEFT JOIN sale_commission scn
            ON scn.saleguid = s.saleguid

        WHERE oit.transaction_identifier_transactionid = %s
    """

    cursor.execute(query, (transactionid,))
    row = cursor.fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Transaction not found")

    columns = [desc[0] for desc in cursor.description]
    data = dict(zip(columns, row))

    skyslope_match = data.get("saleguid") is not None

    return {
        "transactionid": transactionid,
        "otherincome_transactions": {
            "property_address": data.get("property_address"),
            "income_type": data.get("income_type"),
            "income_received_date": data.get("income_received_date"),
            "income_received": data.get("income_received"),
            "gross_commission": data.get("gross_commission"),
            "agent_net": data.get("agent_net"),
            "brokerage_net": data.get("brokerage_net"),
            "agents": data.get("agents"),
            "tags": data.get("tags"),
            "effective_at": data.get("effective_at"),
            "finalized_date": data.get("finalized_date"),
            "status": data.get("transaction_status"),
            "transaction_specialist": data.get("transaction_specialist"),
            "skyslopefileid": data.get("skyslopefileid")
        },
        "skyslope": {
            "match": skyslope_match,
            "saleguid": data.get("saleguid"),
            "property_address": data.get("propertyaddress"),
            "listingprice": data.get("listingprice"),
            "saleprice": data.get("saleprice"),
            "mlsnumber": data.get("mlsnumber"),
            "buying_agent": data.get("skyslope_buying_agent_name"),
            "reviewer_full_name": data.get("reviewer_full_name"),
            "status": data.get("status"),
            "contractacceptancedate": data.get("contractacceptancedate"),
            "escrowclosingdate": data.get("escrowclosingdate"),
            "officegrosscommissiononsale": data.get("officegrosscommissiononsale")
        }
    }

@router.get("/otherincome/noskyslopefileid/download")
def download_otherincome_no_skyslopefileid(conn=Depends(get_db)):
    cursor = conn.cursor()

    query = """
        SELECT *
        FROM otherincome_transactions
        WHERE skyslopefileid IS NULL
        ORDER BY income_received_date DESC NULLS LAST
    """

    cursor.execute(query)
    data = cursor.fetchall()

    columns = [desc[0] for desc in cursor.description]

    rows_to_export = []
    for row in data:
        row_dict = dict(zip(columns, row))
        export_row = {}

        for col in columns:
            val = row_dict.get(col)

            if isinstance(val, Decimal):
                val = float(val)
            elif isinstance(val, (datetime.date, datetime.datetime)):
                val = val.strftime("%Y-%m-%d")
            elif isinstance(val, bool):
                val = "Yes" if val else "No"
            elif val is None:
                val = ""

            export_row[col] = val

        rows_to_export.append(export_row)

    df = pd.DataFrame(rows_to_export, columns=columns)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="No SkySlope File ID", index=False)

        worksheet = writer.sheets["No SkySlope File ID"]

        worksheet.views.sheetView[0].showGridLines = True
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

        font_body = Font(name="Segoe UI", size=10)
        fill_even = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin", color="D0D5DD"),
            right=Side(style="thin", color="D0D5DD"),
            top=Side(style="thin", color="D0D5DD"),
            bottom=Side(style="thin", color="D0D5DD")
        )

        worksheet.row_dimensions[1].height = 28
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_header
            cell.border = thin_border

        currency_cols = []
        date_cols = []
        center_cols = []

        currency_keywords = ["price", "commission", "amount", "gross", "net", "income"]
        date_keywords = ["date"]
        center_keywords = ["id", "guid", "status"]

        for idx, col_name in enumerate(df.columns):
            col_name_lower = col_name.lower()
            if any(kw in col_name_lower for kw in currency_keywords):
                currency_cols.append(idx + 1)
            elif any(kw in col_name_lower for kw in date_keywords):
                date_cols.append(idx + 1)
            elif any(kw in col_name_lower for kw in center_keywords):
                center_cols.append(idx + 1)

        for row_num in range(2, len(df) + 2):
            worksheet.row_dimensions[row_num].height = 20
            is_even_row = (row_num % 2 == 0)

            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = font_body
                cell.border = thin_border

                if is_even_row:
                    cell.fill = fill_even

                val = cell.value

                if col_num in currency_cols:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(val, (int, float)):
                        cell.number_format = '$#,##0.00'
                elif col_num in date_cols:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_num in center_cols:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                val = cell.value
                if val is not None:
                    if cell.column in currency_cols and isinstance(val, (int, float)):
                        val_len = len(f"${val:,.2f}")
                    else:
                        val_len = len(str(val))
                    if val_len > max_len:
                        max_len = val_len

            worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    output.seek(0)

    filename = "otherincome_no_skyslopefileid_report.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )