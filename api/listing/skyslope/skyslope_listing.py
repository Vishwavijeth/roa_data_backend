from fastapi import APIRouter, HTTPException, Query, Depends, Response
from db import get_db
from services.loaders import get_skyslope_sync
import io
import datetime
from decimal import Decimal
import pandas as pd
from typing import List, Optional
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from common.pagination import PaginationResponseWithCount, PaginationData
from common.response import FilterResponse

router = APIRouter()


def norm(x):
    return str(x or "").replace("\u00A0", "").strip().lower()


def apply_skyslope_filters(
    base_filter: str,
    params: list,
    from_close_date: Optional[str] = None,
    to_close_date: Optional[str] = None,
    status: Optional[List[str]] = None,
    stage: Optional[List[str]] = None,
    search: Optional[str] = None,
    not_in_be: bool = False,
):
    if status:
        cleaned_status = [s.strip().lower() for s in status if s and s.strip()]
        if cleaned_status:
            placeholders = ", ".join(["%s"] * len(cleaned_status))
            base_filter += f" AND LOWER(TRIM(s.status)) IN ({placeholders})"
            params.extend(cleaned_status)

    if stage:
        cleaned_stage = [s.strip().lower() for s in stage if s and s.strip()]
        if cleaned_stage:
            placeholders = ", ".join(["%s"] * len(cleaned_stage))
            base_filter += f" AND LOWER(TRIM(st.name)) IN ({placeholders})"
            params.extend(cleaned_stage)

    if from_close_date:
        base_filter += " AND s.escrowclosingdate >= %s"
        params.append(from_close_date)

    if to_close_date:
        base_filter += " AND s.escrowclosingdate <= %s"
        params.append(to_close_date)

    if search:
        search_value = f"%{search.lower().strip()}%"
        base_filter += """
            AND (
                EXISTS (
                    SELECT 1
                    FROM brokerage_engine be_search
                    WHERE be_search.skyslopefileid = s.saleguid
                      AND LOWER(be_search.transaction_identifier_transactionid::text) LIKE %s
                )
                OR EXISTS (
                    SELECT 1
                    FROM sale_property sp_search
                    WHERE sp_search.saleguid = s.saleguid
                      AND LOWER(
                        CONCAT_WS(', ',
                            CONCAT_WS(' ', sp_search.streetnumber, sp_search.streetaddress),
                            sp_search.city,
                            sp_search.state,
                            sp_search.zip
                        )
                      ) LIKE %s
                )
                OR LOWER(s.saleguid::text) LIKE %s
            )
        """
        params.extend([search_value, search_value, search_value])

    if not_in_be:
        base_filter += """
            AND NOT EXISTS (
                SELECT 1
                FROM brokerage_engine be2
                WHERE be2.skyslopefileid = s.saleguid
            )
            AND NOT EXISTS (
                SELECT 1
                FROM otherincome_transactions oit
                WHERE oit.skyslopefileid = s.saleguid
            )
        """

    return base_filter, params


@router.get("/skyslope/sync_info")
def skyslope_sync_info(conn=Depends(get_db)):
    sync_info = get_skyslope_sync(conn)
    return {
        "sync_info": sync_info,
    }


@router.get("/skyslope-listing-filters", response_model=FilterResponse)
def skyslope_listing_filters(conn=Depends(get_db)):
    cursor = conn.cursor()

    status_query = """
        SELECT DISTINCT TRIM(status) AS status
        FROM sale
        WHERE status IS NOT NULL
          AND TRIM(status) <> ''
        ORDER BY TRIM(status)
    """
    cursor.execute(status_query)
    status_list = [row[0] for row in cursor.fetchall()]

    stage_query = """
        SELECT DISTINCT TRIM(name) AS name
        FROM stage
        WHERE name IS NOT NULL
          AND TRIM(name) <> ''
        ORDER BY TRIM(name)
    """
    cursor.execute(stage_query)
    stage_list = [row[0] for row in cursor.fetchall()]

    return FilterResponse(
        filters={
            "status_list": status_list,
            "stage_list": stage_list,
            "not_in_be": False,
        }
    )


@router.get("/skyslope-listing", response_model=PaginationResponseWithCount[dict])
def skyslope_api(
    page: int = Query(default=1, ge=1),
    from_close_date: Optional[str] = Query(default=None),
    to_close_date: Optional[str] = Query(default=None),
    status: Optional[List[str]] = Query(default=None),
    stage: Optional[List[str]] = Query(default=None),
    search: Optional[str] = Query(default=None),
    not_in_be: bool = Query(default=False),
    conn=Depends(get_db)
):
    cursor = conn.cursor()

    limit = 50
    offset = (page - 1) * limit

    count_base_filter = """
        FROM sale s
        LEFT JOIN sale_property sp ON s.saleguid = sp.saleguid
        LEFT JOIN brokerage_engine be ON be.skyslopefileid = s.saleguid
        LEFT JOIN stage st ON s.stageid = st.stageid
        WHERE 1=1
    """
    count_params = []

    count_base_filter, count_params = apply_skyslope_filters(
        base_filter=count_base_filter,
        params=count_params,
        from_close_date=from_close_date,
        to_close_date=to_close_date,
        status=status,
        stage=stage,
        search=search,
        not_in_be=not_in_be,
    )

    count_query = "SELECT COUNT(DISTINCT s.saleguid) " + count_base_filter
    cursor.execute(count_query, count_params)
    total_count = cursor.fetchone()[0]

    data_query = """
        SELECT
            s.saleguid,
            CONCAT_WS(', ',
                CONCAT_WS(' ', sp.streetnumber, sp.streetaddress),
                sp.city,
                sp.state,
                sp.zip
            ) AS propertyaddress,
            s.escrowclosingdate AS close_date,
            TRIM(s.status) AS status,
            TRIM(st.name) AS stage_name,
            COALESCE(
                be.transaction_identifier_transactionid::text,
                (
                    SELECT oit.transaction_identifier_transactionid::text
                    FROM otherincome_transactions oit
                    WHERE oit.skyslopefileid = s.saleguid
                      AND oit.transaction_identifier_transactionid IS NOT NULL
                    ORDER BY oit.finalized_date DESC NULLS LAST,
                             oit.income_received_date DESC NULLS LAST
                    LIMIT 1
                ),
                'No related transaction data'
            ) AS transaction_id,
            NULLIF(
                COALESCE(
                    (
                        SELECT STRING_AGG(
                            TRIM(
                                COALESCE(sc.firstname, '') || ' ' ||
                                COALESCE(sc.lastname, '')
                            ),
                            ', '
                        )
                        FROM sale_contact sc
                        WHERE sc.saleguid = s.saleguid
                          AND LOWER(sc.role) = 'buyer'
                    ),
                    ''
                ),
                ''
            ) AS buyer_name,
            NULLIF(
                (
                    SELECT TRIM(COALESCE(u.firstname, '') || ' ' || COALESCE(u.lastname, ''))
                    FROM users u
                    WHERE u.userguid = s.agentguid
                ),
                ''
            ) AS buyer_agent_name,
            NULLIF(
                TRIM(COALESCE(r.firstname, '') || ' ' || COALESCE(r.lastname, '')),
                ''
            ) AS reviewer
        FROM sale s
        LEFT JOIN users r ON s.reviewerguid = r.userguid
        LEFT JOIN sale_property sp ON s.saleguid = sp.saleguid
        LEFT JOIN brokerage_engine be ON be.skyslopefileid = s.saleguid
        LEFT JOIN stage st ON s.stageid = st.stageid
        WHERE 1=1
    """
    data_params = []

    data_query, data_params = apply_skyslope_filters(
        base_filter=data_query,
        params=data_params,
        from_close_date=from_close_date,
        to_close_date=to_close_date,
        status=status,
        stage=stage,
        search=search,
        not_in_be=not_in_be,
    )

    data_query += """
        ORDER BY s.escrowclosingdate DESC NULLS LAST
        LIMIT %s OFFSET %s
    """
    cursor.execute(data_query, data_params + [limit, offset])

    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()
    data = [dict(zip(columns, row)) for row in rows]

    total_pages = (total_count + limit - 1) // limit if total_count else 1
    has_next = page < total_pages

    return PaginationResponseWithCount[dict](
        data=PaginationData[dict](
            total_count=total_count,
            items=data
        ),
        page=page,
        page_size=limit,
        count=len(data),
        total_pages=total_pages,
        has_next=has_next,
    )


@router.get("/skyslope/detail")
def skyslope_detail(saleguid: str, conn=Depends(get_db)):
    cursor = conn.cursor()

    skyslope_query = """
        SELECT
            s.saleguid,
            s.saleguid AS skyslope_saleguid,
            CONCAT_WS(', ',
                CONCAT_WS(' ', sp.streetnumber, sp.streetaddress),
                sp.city,
                sp.state,
                sp.zip
            ) AS propertyaddress,
            s.listingprice,
            s.saleprice,
            s.mlsnumber,
            s.contractacceptancedate,
            s.escrowclosingdate,
            TRIM(s.status) AS status,
            TRIM(st.name) AS stage_name,
            COALESCE(u.firstname || ' ' || u.lastname, NULL) AS buyer_agent_name,
            COALESCE(u.email, NULL) AS buyer_agent_email,
            COALESCE(r.firstname || ' ' || r.lastname, NULL) AS reviewer,
            COALESCE(
                (
                    SELECT STRING_AGG(
                        TRIM(
                            COALESCE(sc.firstname, '') || ' ' ||
                            COALESCE(sc.lastname, '')
                        ),
                        ', '
                    )
                    FROM sale_contact sc
                    WHERE sc.saleguid = s.saleguid
                      AND LOWER(sc.role) = 'buyer'
                ),
                NULL
            ) AS buyer,
            COALESCE(
                (
                    SELECT STRING_AGG(
                        TRIM(
                            COALESCE(sc.firstname, '') || ' ' ||
                            COALESCE(sc.lastname, '')
                        ),
                        ', '
                    )
                    FROM sale_contact sc
                    WHERE sc.saleguid = s.saleguid
                      AND LOWER(sc.role) = 'seller'
                ),
                NULL
            ) AS seller
        FROM sale s
        LEFT JOIN users u ON s.agentguid = u.userguid
        LEFT JOIN users r ON s.reviewerguid = r.userguid
        LEFT JOIN sale_property sp ON s.saleguid = sp.saleguid
        LEFT JOIN stage st ON s.stageid = st.stageid
        WHERE s.saleguid = %s
    """

    cursor.execute(skyslope_query, [saleguid])
    row = cursor.fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Sale not found")

    columns = [desc[0] for desc in cursor.description]
    skyslope_data = dict(zip(columns, row))

    be_query = """
        SELECT
            be.transaction_identifier_transactionid::text AS transactionid,
            be.property_address,
            be.sale_price,
            be.listing_price,
            be.listing_office AS office,
            be.buyer_name,
            be.seller_name,
            be.buying_agent_name,
            be.contract_date,
            be.closed_date,
            be.tags,
            be.transaction_status,
            be.transaction_specialist,
            be.skyslopefileid,
            be.total_gross_commission,
            CASE
                WHEN LOWER(be.transaction_status) = 'cancelled'
                    THEN 'ignored_cancelled_duplicate'
                ELSE 'active'
            END AS record_role
        FROM brokerage_engine be
        WHERE be.skyslopefileid = %s
        ORDER BY
            CASE
                WHEN LOWER(be.transaction_status) = 'cancelled' THEN 2
                ELSE 1
            END,
            be.closed_date DESC NULLS LAST
    """

    cursor.execute(be_query, [saleguid])
    be_rows = cursor.fetchall()
    be_columns = [desc[0] for desc in cursor.description]
    brokerage_engine_records = [dict(zip(be_columns, r)) for r in be_rows]

    other_income_query = """
        SELECT
            oit.transaction_identifier_transactionid::text AS transactionid,
            oit.property_address,
            oit.income_type,
            oit.income_received_date,
            oit.income_received,
            oit.gross_commission,
            oit.agent_net,
            oit.brokerage_net,
            oit.agents,
            oit.client_type,
            oit.client_name,
            oit.client_phone,
            oit.client_email,
            oit.tags,
            oit.effective_at,
            oit.finalized_date,
            oit.transaction_specialist,
            oit.transaction_status,
            oit.skyslopefileid,
            'other_income' AS record_role
        FROM otherincome_transactions oit
        WHERE oit.skyslopefileid = %s
        ORDER BY oit.finalized_date DESC NULLS LAST, oit.income_received_date DESC NULLS LAST
    """

    cursor.execute(other_income_query, [saleguid])
    oi_rows = cursor.fetchall()
    oi_columns = [desc[0] for desc in cursor.description]
    otherincome_records = [dict(zip(oi_columns, r)) for r in oi_rows]

    return {
        "saleguid": skyslope_data["saleguid"],
        "skyslope": {
            "saleguid": skyslope_data["skyslope_saleguid"],
            "propertyaddress": skyslope_data["propertyaddress"],
            "listingprice": skyslope_data["listingprice"],
            "saleprice": skyslope_data["saleprice"],
            "mlsnumber": skyslope_data["mlsnumber"],
            "seller": skyslope_data["seller"],
            "buyer": skyslope_data["buyer"],
            "buyer_agent": skyslope_data["buyer_agent_name"],
            "buyer_agent_email": skyslope_data["buyer_agent_email"],
            "reviewer": skyslope_data["reviewer"],
            "status": skyslope_data["status"],
            "stage": skyslope_data["stage_name"],
            "contractacceptancedate": skyslope_data["contractacceptancedate"],
            "escrowclosingdate": skyslope_data["escrowclosingdate"]
        },
        "brokerage_engine_records": brokerage_engine_records,
        "otherincome_transactions": otherincome_records,
    }


@router.get("/skyslope/download")
def skyslope_download(
    from_close_date: Optional[str] = Query(default=None),
    to_close_date: Optional[str] = Query(default=None),
    status: Optional[List[str]] = Query(default=None),
    stage: Optional[List[str]] = Query(default=None),
    search: Optional[str] = Query(default=None),
    not_in_be: bool = Query(default=False),
    conn=Depends(get_db)
):
    cursor = conn.cursor()

    sale_base_filter = """
        FROM sale s
        LEFT JOIN sale_property sp ON s.saleguid = sp.saleguid
        LEFT JOIN stage st ON s.stageid = st.stageid
        LEFT JOIN sale_commission sc2 ON sc2.saleguid = s.saleguid
        WHERE 1=1
    """

    sale_params = []

    sale_base_filter, sale_params = apply_skyslope_filters(
        base_filter=sale_base_filter,
        params=sale_params,
        from_close_date=from_close_date,
        to_close_date=to_close_date,
        status=status,
        stage=stage,
        search=search,
        not_in_be=not_in_be,
    )

    sale_query = f"""
        SELECT
            'sale' AS record_source,
            s.saleguid AS saleguid,
            NULL::text AS transactionid,
            CONCAT_WS(', ',
                CONCAT_WS(' ', sp.streetnumber, sp.streetaddress),
                sp.city,
                sp.state,
                sp.zip
            ) AS property_address,
            NULLIF(
                (
                    SELECT STRING_AGG(
                        TRIM(COALESCE(sc.firstname, '') || ' ' || COALESCE(sc.lastname, '')),
                        ', '
                        ORDER BY sc.firstname, sc.lastname
                    )
                    FROM sale_contact sc
                    WHERE sc.saleguid = s.saleguid
                      AND LOWER(sc.role) = 'buyer'
                ),
                ''
            ) AS buyer_name,
            NULLIF(
                (
                    SELECT STRING_AGG(
                        TRIM(COALESCE(sc.firstname, '') || ' ' || COALESCE(sc.lastname, '')),
                        ', '
                        ORDER BY sc.firstname, sc.lastname
                    )
                    FROM sale_contact sc
                    WHERE sc.saleguid = s.saleguid
                      AND LOWER(sc.role) = 'seller'
                ),
                ''
            ) AS seller_name,
            s.saleprice AS sale_price,
            s.listingprice AS listing_price,
            s.escrowclosingdate AS escrow_close_date,
            TRIM(s.status) AS status,
            TRIM(st.name) AS stage,
            s.dealtype AS dealtype,
            sc2."officegrosscommissiononsale" AS office_gross_commission_on_sale,
            sc2."salecommissionamount" AS sale_commission_amount,
            sc2."listingcommissionamount" AS listing_commission_amount,
            NULL::text AS income_type,
            NULL::date AS income_received_date,
            NULL::numeric AS income_received,
            NULL::numeric AS gross_commission,
            NULL::numeric AS agent_net,
            NULL::numeric AS brokerage_net,
            NULL::text AS client_type,
            NULL::text AS client_name,
            NULL::text AS client_email,
            NULL::text AS transaction_status,
            NULL::text AS transaction_specialist
        {sale_base_filter}
    """

    other_income_query = """
        SELECT
            'other_income' AS record_source,
            oit.skyslopefileid AS saleguid,
            oit.transaction_identifier_transactionid::text AS transactionid,
            oit.property_address,
            NULL::text AS buyer_name,
            oit.client_name AS seller_name,
            NULL::numeric AS sale_price,
            NULL::numeric AS listing_price,
            oit.finalized_date AS escrow_close_date,
            NULL::text AS status,
            NULL::text AS stage,
            NULL::text AS dealtype,
            NULL::numeric AS office_gross_commission_on_sale,
            NULL::numeric AS sale_commission_amount,
            NULL::numeric AS listing_commission_amount,
            oit.income_type,
            oit.income_received_date,
            oit.income_received,
            oit.gross_commission,
            oit.agent_net,
            oit.brokerage_net,
            oit.client_type,
            oit.client_name,
            oit.client_email,
            oit.transaction_status,
            oit.transaction_specialist
        FROM otherincome_transactions oit
        WHERE 1=1
    """

    oi_params = []

    if search:
        search_value = f"%{search.lower().strip()}%"
        other_income_query += """
            AND (
                LOWER(COALESCE(oit.transaction_identifier_transactionid::text, '')) LIKE %s
                OR LOWER(COALESCE(oit.property_address, '')) LIKE %s
                OR LOWER(COALESCE(oit.client_name, '')) LIKE %s
                OR LOWER(COALESCE(oit.client_email, '')) LIKE %s
                OR LOWER(COALESCE(oit.skyslopefileid::text, '')) LIKE %s
            )
        """
        oi_params.extend([search_value, search_value, search_value, search_value, search_value])

    if not_in_be:
        other_income_query += """
            AND NOT EXISTS (
                SELECT 1
                FROM sale s
                WHERE s.saleguid = oit.skyslopefileid
            )
        """

    final_query = f"""
        {sale_query}
        UNION ALL
        {other_income_query}
        ORDER BY escrow_close_date DESC NULLS LAST, saleguid, transactionid
    """

    cursor.execute(final_query, sale_params + oi_params)
    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()
    data = [dict(zip(columns, row)) for row in rows]

    columns_map = {
        "record_source": "Record Source",
        "saleguid": "Sale GUID",
        "transactionid": "Transaction ID",
        "property_address": "Property Address",
        "buyer_name": "Buyer Name",
        "seller_name": "Seller Name",
        "sale_price": "Sale Price",
        "listing_price": "Listing Price",
        "escrow_close_date": "Escrow Close Date",
        "status": "Status",
        "stage": "Stage",
        "dealtype": "Deal Type",
        "office_gross_commission_on_sale": "Office Gross Commission On Sale",
        "sale_commission_amount": "Sale Commission Amount",
        "listing_commission_amount": "Listing Commission Amount",
        "income_type": "Income Type",
        "income_received_date": "Income Received Date",
        "income_received": "Income Received",
        "gross_commission": "Gross Commission",
        "agent_net": "Agent Net",
        "brokerage_net": "Brokerage Net",
        "client_type": "Client Type",
        "client_name": "Client Name",
        "client_email": "Client Email",
        "transaction_status": "Transaction Status",
        "transaction_specialist": "Transaction Specialist",
    }

    rows_to_export = []
    for record in data:
        row_dict = {}
        for key, header in columns_map.items():
            val = record.get(key)

            if isinstance(val, Decimal):
                val = float(val)
            elif isinstance(val, (datetime.date, datetime.datetime)):
                val = val.strftime("%Y-%m-%d")
            elif isinstance(val, bool):
                val = "Yes" if val else "No"
            elif val is None:
                val = ""

            row_dict[header] = val

        rows_to_export.append(row_dict)

    df = pd.DataFrame(rows_to_export)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="SkySlope Report", index=False)

        worksheet = writer.sheets["SkySlope Report"]
        worksheet.freeze_panes = "A2"

        font_header = Font(name="Segoe UI", size=11, bold=True)
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        font_body = Font(name="Segoe UI", size=10)

        worksheet.row_dimensions[1].height = 28

        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = font_header
            cell.alignment = align_header

        currency_cols = []
        date_cols = []
        center_cols = []

        currency_keywords = [
            "sale price",
            "listing price",
            "commission",
            "income received",
            "gross commission",
            "agent net",
            "brokerage net",
        ]
        date_keywords = ["date"]
        center_keywords = [
            "record source",
            "sale guid",
            "transaction id",
            "status",
            "stage",
            "deal type",
            "client type",
            "transaction status",
            "transaction specialist",
        ]

        for idx, col_name in enumerate(df.columns):
            col_name_lower = col_name.lower()
            if any(kw in col_name_lower for kw in currency_keywords):
                currency_cols.append(idx + 1)
            elif any(kw in col_name_lower for kw in date_keywords):
                date_cols.append(idx + 1)
            elif any(kw in col_name_lower for kw in center_keywords):
                center_cols.append(idx + 1)

        for row_num in range(2, len(df) + 2):
            worksheet.row_dimensions[row_num].height = 20

            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = font_body
                val = cell.value

                if col_num in currency_cols:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(val, (int, float)):
                        cell.number_format = '$#,##0.00'
                elif col_num in date_cols:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_num in center_cols:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                val = cell.value
                if val is not None:
                    if cell.column in currency_cols and isinstance(val, (int, float)):
                        val_len = len(f"${val:,.2f}")
                    else:
                        val_len = len(str(val))
                    if val_len > max_len:
                        max_len = val_len

            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output.seek(0)

    filename = "skyslope_sale_report.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )